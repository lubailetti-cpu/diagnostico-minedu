"""
Genera Excel comparativo entre cuestionario original de Javier (33 preguntas)
y la version actual del diagnostico Moderniza (25 preguntas).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Estados y colores
ESTADOS = {
    "MANTENIDA": "FFFFFF",            # blanco
    "REFORMULADA": "FEF9EC",          # amarillo suave
    "REFORMATEADA": "FEF9EC",         # amarillo suave
    "MOVIDA": "EFF4FF",               # azul soft
    "FUSIONADA": "ECFDF5",            # verde soft
    "ELIMINADA": "FFF1F2",            # rosa
    "NUEVA": "F3EEFF",                # morado soft
}

# Comparativo: cada fila es una pregunta de Javier o una nueva
# Campos: javier_num, javier_seccion, javier_texto, estado, actual_num, actual_seccion, actual_texto, justificacion
COMPARATIVO = [
    # ============ A. PERFIL ============
    ("Inicio", "Encabezado", "Este formulario registrará su nombre, escriba su nombre (obligatorio)", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Pedido de Javier. El campo de nombre se quita para garantizar anonimato real. Esto es consistente con la promesa que la encuesta original ya hacía en su intro pero contradecía al pedir nombre obligatorio."),

    ("1", "A. Perfil", "¿A qué régimen laboral pertenece?", "REFORMATEADA",
     "1", "A. Perfil", "¿A qué régimen laboral pertenece? (desplegable)",
     "Convertida a select desplegable para ahorrar espacio en pantalla. Las 6 opciones se mantienen idénticas. Pedido de Lucía: 'P1, P3, P4 quiero que sean desplegables'."),

    ("2", "A. Perfil", "¿A qué unidad organizacional (U.O.) del Minedu pertenece?", "MANTENIDA",
     "2", "A. Perfil", "¿A qué unidad organizacional (U.O.) del Minedu pertenece? (desplegable)",
     "Mantenida tal cual. Lista completa de unidades MINEDU intacta."),

    ("3", "A. Perfil", "¿Cuál es el rol principal que ejerce actualmente dentro de su U.O.?", "REFORMATEADA",
     "3", "A. Perfil", "¿Cuál es el rol principal que ejerce actualmente dentro de su U.O.? (desplegable)",
     "Convertida a select desplegable. Las 5 opciones se mantienen."),

    ("4", "A. Perfil", "En general, ¿cuántos años de experiencia tiene en el sector público?", "REFORMATEADA",
     "4", "A. Perfil", "En general, ¿cuántos años de experiencia tiene en el sector público? (desplegable)",
     "Convertida a select desplegable. Los 4 rangos se mantienen."),

    # ============ B. DISTRIBUCIÓN DEL TIEMPO ============
    ("5", "B. Distribución", "¿Con qué frecuencia elabora o revisa instrumentos de gestión pública?", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada porque la matriz de volumen por tipo de documento (P6 actual) ya captura esta información con más detalle: muestra exactamente cuántos documentos de cada tipo y con qué frecuencia."),

    ("6", "B. Distribución", "¿Cuántos documentos o productos documentales elabora o revisa en promedio al mes?", "REFORMULADA",
     "6", "B. Tu trabajo documental", "Para cada tipo de documento, indique cuántos elabora o revisa al mes y cuánto tiempo le toma en promedio elaborar cada uno (matriz por 9 tipos)",
     "Reformulada como matriz por TIPO de documento (ficha proceso, diagrama, informe, etc.). Razón: una sola cifra global no permite calcular impacto de Moderniza por tipo. Esta matriz es el núcleo de la línea de base: permite decir 'antes 4h por ficha, ahora 1.2h' después del piloto."),

    ("7", "B. Distribución", "% tiempo FUNCIONES SUSTANTIVAS (campo numérico 0-100)", "REFORMATEADA",
     "5", "B. Tu trabajo documental", "Distribución de su tiempo laboral semanal (barra segmentada con 3 puntos arrastrables, 4 colores)",
     "Las 4 preguntas separadas de % (sustantivas / documentos / administrativas / externas) se consolidaron en UNA sola barra segmentada interactiva. Razón: los campos numéricos sueltos generaban errores frecuentes (la gente no llega a sumar 100). La barra visual fuerza que sume 100 automáticamente y es más rápido de llenar."),

    ("8", "B. Distribución", "% tiempo ELABORACIÓN/REVISIÓN documentos normativos", "FUSIONADA",
     "5", "B. Tu trabajo documental", "(integrada en la barra segmentada de distribución)",
     "Consolidada con las otras 3 preguntas de % en una sola barra segmentada."),

    ("9", "B. Distribución", "% tiempo TAREAS ADMINISTRATIVAS REPETITIVAS", "FUSIONADA",
     "5", "B. Tu trabajo documental", "(integrada en la barra segmentada de distribución)",
     "Consolidada con las otras 3 preguntas de % en una sola barra segmentada."),

    ("10", "B. Distribución", "% tiempo ATENCIÓN DE DEMANDAS EXTERNAS NO PLANIFICADAS", "FUSIONADA",
     "5", "B. Tu trabajo documental", "(integrada en la barra segmentada de distribución)",
     "Consolidada con las otras 3 preguntas de % en una sola barra segmentada."),

    ("11", "B. Distribución", "¿Con qué frecuencia recibe solicitudes urgentes externas?", "MOVIDA",
     "14", "E. Ritmo y demandas externas", "¿Con qué frecuencia recibe solicitudes urgentes de entidades externas (Congreso, Contraloría, PCM, Defensoría, GOREs u otras)?",
     "Movida a su propia sección 'Ritmo y demandas externas' para tener mejor agrupación lógica. Texto mantenido."),

    ("12", "B. Distribución", "Cuando recibe estas solicitudes externas, ¿cuál es el plazo habitual?", "MOVIDA",
     "15", "E. Ritmo y demandas externas", "Cuando recibe estas solicitudes externas, ¿cuál es el plazo habitual que le otorgan para responder?",
     "Movida a 'Ritmo y demandas externas'. Texto mantenido."),

    ("13", "B. Distribución", "¿Con qué frecuencia tiene que realizar reprocesos en un mismo instrumento?", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada por pedido de Lucía. Razón: se superpone con el % de devolución (P9 actual) y las rondas de revisión (P10 actual), que son métricas más concretas y duras de manipular."),

    ("14", "B. Distribución", "Cuando un documento requiere reproceso, ¿cuántas rondas de revisión/devolución se producen?", "MOVIDA",
     "10", "C. Reproceso y calidad", "Cuando un documento requiere reproceso, ¿cuántas rondas de revisión/devolución se producen en promedio antes de su aprobación final?",
     "Movida a sección C 'Reproceso y calidad' (nueva sección que agrupa todas las métricas de calidad). Texto mantenido."),

    ("15", "B. Distribución", "¿Con qué frecuencia los criterios o instrucciones para elaborar un documento cambian?", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada por pedido de Lucía. Se sentía duplicada con la P13 original (frecuencia reprocesos) y con la P11 actual (razones de devolución, donde una opción es 'falta de alineamiento normativo')."),

    ("16", "B. Distribución", "En el último trimestre, ¿la carga ha ocasionado retrasos en POI?", "REFORMULADA",
     "16", "F. Oportunidades en tu área", "Afirmación (escala 1-4): 'La carga documental y administrativa ha ocasionado retrasos en el cumplimiento de actividades planificadas en el POI de mi unidad'",
     "Reformulada como AFIRMACIÓN en escala semáforo 1-4 dentro del bloque de oportunidades. Razón: alinea con el formato del resto de afirmaciones del bloque BCP. Es la primera afirmación del grupo. Pedido de Lucía."),

    ("17", "B. Distribución", "En el último mes, ¿cuántas veces ha tenido que extender su jornada laboral?", "MOVIDA",
     "17", "F. Oportunidades en tu área", "En el último mes, ¿cuántas veces ha tenido que extender su jornada laboral para cumplir con entregables documentales o administrativos?",
     "Movida a sección F 'Oportunidades en tu área' antes de 'a qué destinaría el tiempo'. Texto mantenido. Pedido de Lucía."),

    ("18", "B. Distribución", "Cuando elabora un instrumento nuevo desde cero, ¿cuánto tiempo le toma?", "FUSIONADA",
     "6", "B. Tu trabajo documental", "(integrada en la matriz de volumen + tiempo por tipo de documento)",
     "Fusionada con la nueva matriz por tipo. Razón: la pregunta original era demasiado genérica ('un instrumento nuevo'). Ahora se mide por TIPO: cuánto tiempo toma una ficha de proceso, un diagrama, un informe, etc. Esto permite medir impacto de Moderniza con precisión por tipo."),

    ("19", "B. Distribución", "¿Cuánto tiempo dedica a buscar normativa antes de elaborar un instrumento?", "MOVIDA",
     "12", "D. Búsqueda de información", "En promedio, ¿cuánto tiempo dedica a buscar normativa, guías metodológicas o modelos de referencia ANTES de empezar a elaborar un instrumento?",
     "Movida a una nueva sección D 'Búsqueda de información'. Texto mantenido."),

    # ============ C. INSTRUMENTOS QUE MÁS CARGA GENERAN ============
    ("20", "C. Instrumentos", "¿Cuáles son los 3 documentos o productos que más tiempo de elaboración le demandan?", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada porque la matriz de volumen + tiempo (P6 actual) ya identifica de manera más rigurosa cuáles son los documentos que más carga generan: con el cruce volumen × tiempo unitario se obtiene la carga real total."),

    ("21", "C. Instrumentos", "¿Cuál es la principal dificultad al elaborar estos instrumentos?", "MANTENIDA",
     "7", "B. Tu trabajo documental", "¿Cuál es la principal dificultad al elaborar estos instrumentos? (max 2 opciones)",
     "Mantenida con las 8 opciones originales. Solo movida a la sección de trabajo documental."),

    ("22", "C. Instrumentos", "¿Tiene acceso a plantillas, guías o herramientas estandarizadas?", "MANTENIDA",
     "8", "B. Tu trabajo documental", "¿Tiene acceso a plantillas, guías o herramientas estandarizadas para elaborar estos instrumentos?",
     "Mantenida con las 4 opciones originales."),

    ("23", "C. Instrumentos", "¿Considera que la carga de trabajo en su equipo está distribuida equitativamente?", "REFORMULADA",
     "16", "F. Oportunidades en tu área", "Afirmación: 'La carga de trabajo de mi equipo NO está distribuida de la mejor manera' (escala semáforo 1-4)",
     "Reformulada como AFIRMACIÓN en negativo (consistente con el resto del bloque BCP). Pedido de Lucía: redactar en negativo y meter en oportunidades."),

    ("24", "C. Instrumentos", "Si pudiera recuperar tiempo de tareas repetitivas, ¿a qué actividad lo destinaría?", "MOVIDA",
     "18", "F. Oportunidades en tu área", "Si pudiera recuperar tiempo de las tareas repetitivas y administrativas, ¿a qué actividades lo destinaría? (max 2)",
     "Movida a sección F 'Oportunidades'. Cambiada a checkboxes con max 2 opciones (antes era una sola elección). Pedido de Lucía: 'P25 a qué actividades lo destinaría — déjala con max 2'."),

    # ============ D. PERCEPCIÓN SOBRE IA ============
    ("25", "D. Percepción IA", "¿Ha utilizado alguna herramienta de IA generativa?", "REFORMULADA",
     "19", "G. Uso actual de inteligencia artificial", "¿Cuál es tu nivel actual de uso de inteligencia artificial? (5 niveles: no uso / básico / pragmático / constructor / avanzado)",
     "Reformulada con la escala de madurez del cuestionario del BCP. Razón: la pregunta original solo distinguía 'regular vs esporádico'. La escala BCP de 4 niveles + 'no uso' segmenta mucho mejor a la población y permite identificar early adopters."),

    ("26", "D. Percepción IA", "Si utiliza herramientas de IA, ¿con qué recursos lo hace? (MINEDU / propias / ambas)", "REFORMULADA",
     "20", "G. Uso actual de inteligencia artificial", "¿Usas alguna herramienta de inteligencia artificial para facilitar tu trabajo? (Sí pagada por mí / Sí gratuita / No uso ninguna)",
     "Reformulada por pedido de Lucía. Razón: el MINEDU no provee licencias de IA a sus servidores, así que la categoría 'provistas por MINEDU' del original no aplica. Se simplifica a 'pagada por mí / gratuita / no uso'."),

    ("Nueva", "D. Percepción IA", "(no existía)", "NUEVA",
     "21", "G. Uso actual de inteligencia artificial", "¿Para qué tareas de tu trabajo usas inteligencia artificial hoy? (max 3) — condicional: solo aparece si en P20 dice que sí usa IA",
     "Pregunta nueva del modelo BCP. Razón: identifica casos de uso reales de IA en el sector público, qué tareas ya se delegan, qué otras podrían sumarse a un workflow tipo Moderniza."),

    ("27", "D. Percepción IA", "Valore su nivel de acuerdo con afirmaciones sobre la plataforma (6 afirmaciones, escala 1-5)", "REFORMULADA",
     "22", "H. Percepción sobre la plataforma", "Valore su nivel de acuerdo con afirmaciones sobre una posible plataforma (4 afirmaciones, escala 1-5)",
     "Reducida de 6 a 4 afirmaciones. Pedido de Lucía: quitar 'Confiaría en los borradores si pudiera revisarlos antes' y 'Una IA con errores frecuentes me haría dejar de usarla'. Razón: la primera es obvia (cualquiera diría que sí), la segunda es prácticamente una amenaza."),

    ("28", "D. Percepción IA", "¿Cuáles serían sus principales preocupaciones al usar IA? (max 2)", "MANTENIDA",
     "23", "H. Percepción sobre la plataforma", "¿Cuáles serían sus principales preocupaciones al usar una herramienta de IA para elaborar instrumentos de gestión? (max 2)",
     "Mantenida con las 8 opciones originales."),

    ("29", "D. Percepción IA", "¿Para cuáles documentos le sería MÁS útil una herramienta de asistencia con IA?", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada por pedido de Lucía. Razón: se duplicaba con la P21 actual (para qué tareas usas IA hoy)."),

    ("30", "D. Percepción IA", "¿Qué condición considera indispensable para adoptar IA en su trabajo?", "MANTENIDA",
     "24", "H. Percepción sobre la plataforma", "¿Qué condición considera indispensable para adoptar una herramienta de IA en su trabajo? (max 2)",
     "Mantenida con las 6 opciones originales."),

    # ============ E. PREGUNTAS ABIERTAS ============
    ("31", "E. Abiertas", "¿Cuáles son las 2 o 3 tareas repetitivas que más tiempo le quitan?", "MANTENIDA",
     "25", "I. Tu voz cuenta", "En sus propias palabras: ¿cuáles son las 2 o 3 tareas repetitivas que más tiempo le quitan en su trabajo semanal?",
     "Mantenida tal cual."),

    ("32", "E. Abiertas", "¿Qué funcionalidad le parecería más útil en una plataforma de asistencia con IA?", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada por pedido de Lucía. Razón: aporta poco valor incremental cuando ya se preguntó qué tareas repetitivas quitan tiempo (P25 actual). Las respuestas terminarían siendo redundantes."),

    ("33", "E. Abiertas", "¿Carga distribuida equitativamente en su equipo? (libre)", "ELIMINADA",
     "—", "—", "(eliminada)",
     "Eliminada por duplicación. La pregunta 23 ya cubría carga equitativa en escala, y la versión actual la integra como afirmación BCP."),

    # ============ NUEVAS QUE NO ESTABAN EN JAVIER ============
    ("Nueva", "C. Reproceso y calidad", "(no existía)", "NUEVA",
     "9", "C. Reproceso y calidad", "De cada 10 documentos que entrega, ¿cuántos le devuelven con observaciones?",
     "Pregunta nueva para medir la TASA REAL de devolución. Razón: junto con las razones de devolución (P11) permite calcular si Moderniza reduce los retrabajos. Esta métrica es central para la línea de base."),

    ("Nueva", "C. Reproceso y calidad", "(no existía)", "NUEVA",
     "11", "C. Reproceso y calidad", "Cuando le devuelven un documento, ¿por qué suele ser? (max 2)",
     "Pregunta nueva. Razón: distingue entre errores de FORMA (que Moderniza ataca directo) y errores de FONDO (que no). Permite atribuir bien el impacto de la plataforma."),

    ("Nueva", "D. Búsqueda de información", "(no existía)", "NUEVA",
     "13", "D. Búsqueda de información", "¿Qué tan trabajoso le resulta asegurarse de que la normativa, guía o modelo que está usando es la versión vigente? (escala 1-4 colores semáforo)",
     "Pregunta nueva. Razón: dimensiona el problema central que ataca un RAG (consulta normativa). Si la mayoría responde 'muy trabajoso', justifica con datos el esfuerzo de armar la base de normas indexadas."),

    ("Nueva", "F. Oportunidades", "(no existía)", "NUEVA",
     "16", "F. Oportunidades en tu área", "Me siento agotado/a al terminar mi semana laboral (escala 1-4 semáforo)",
     "Afirmación nueva en negativo. Razón: mide impacto humano del trabajo documental excesivo. Permite contar la historia de cómo Moderniza mejora bienestar, no solo eficiencia."),

    ("Nueva", "F. Oportunidades", "(no existía)", "NUEVA",
     "16", "F. Oportunidades en tu área", "Me siento poco satisfecho/a con la calidad de los entregables que terminamos produciendo como área (escala 1-4 semáforo)",
     "Afirmación nueva en negativo. Razón: mide percepción de calidad subjetiva. Si Moderniza ayuda a producir mejores documentos, esta métrica debería mejorar."),

    ("Nueva", "F. Oportunidades", "(no existía)", "NUEVA",
     "16", "F. Oportunidades en tu área", "Hay mucho trabajo manual repetitivo en mi área que podría automatizarse (escala 1-4 semáforo)",
     "Afirmación nueva del modelo BCP. Razón: identifica apetito por automatización en cada área. Las áreas con score alto son las que más se beneficiarán de Moderniza."),

    ("Nueva", "F. Oportunidades", "(no existía)", "NUEVA",
     "16", "F. Oportunidades en tu área", "Mucha información importante de mi área está en la cabeza de pocas personas (escala 1-4 semáforo)",
     "Afirmación nueva del modelo BCP. Razón: identifica oportunidad para workflows de gestión del conocimiento (más allá de Moderniza)."),

    ("Nueva", "F. Oportunidades", "(no existía)", "NUEVA",
     "16", "F. Oportunidades en tu área", "Tenemos datos importantes pero dispersos en distintos formatos que nos cuesta consolidar (escala 1-4 semáforo)",
     "Afirmación nueva del modelo BCP. Razón: identifica oportunidad para workflows de consolidación de datos."),

    ("Nueva", "F. Oportunidades", "(no existía)", "NUEVA",
     "16", "F. Oportunidades en tu área", "Nos cuesta transformar la información disponible en indicadores para tomar decisiones (escala 1-4 semáforo)",
     "Afirmación nueva del modelo BCP. Razón: identifica oportunidad para workflows analíticos (dashboards, indicadores)."),
]


def main():
    wb = openpyxl.Workbook()

    # ===== HOJA 1: COMPARATIVO =====
    ws = wb.active
    ws.title = "Comparativo"

    # Encabezados
    headers = [
        "# Javier", "Sección Javier", "Pregunta original Javier",
        "Estado del cambio",
        "# Actual", "Sección actual", "Pregunta actual",
        "Justificación del ajuste"
    ]
    bold_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin = Side(border_style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Filas de datos
    for i, (jn, jsec, jtxt, estado, an, asec, atxt, just) in enumerate(COMPARATIVO, 2):
        row_color = ESTADOS.get(estado, "FFFFFF")
        row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

        values = [jn, jsec, jtxt, estado, an, asec, atxt, just]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = left_wrap
            cell.fill = row_fill
            cell.border = border
            if col == 4:  # columna estado
                cell.font = Font(name='Calibri', size=10, bold=True)

    # Anchos de columna
    widths = [10, 22, 50, 16, 10, 28, 50, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Alto de filas
    for row in range(2, len(COMPARATIVO) + 2):
        ws.row_dimensions[row].height = 60

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # ===== HOJA 2: RESUMEN =====
    ws2 = wb.create_sheet("Resumen")

    # Estadísticas
    total_javier = sum(1 for r in COMPARATIVO if r[0] not in ("Nueva", "Inicio"))
    eliminadas = sum(1 for r in COMPARATIVO if r[3] == "ELIMINADA" and r[0] not in ("Nueva",))
    mantenidas = sum(1 for r in COMPARATIVO if r[3] == "MANTENIDA")
    reformuladas = sum(1 for r in COMPARATIVO if r[3] in ("REFORMULADA", "REFORMATEADA"))
    movidas = sum(1 for r in COMPARATIVO if r[3] == "MOVIDA")
    fusionadas = sum(1 for r in COMPARATIVO if r[3] == "FUSIONADA")
    nuevas = sum(1 for r in COMPARATIVO if r[3] == "NUEVA")

    resumen_rows = [
        ("Métrica", "Cantidad"),
        ("Preguntas en cuestionario original de Javier", "33 + 1 (nombre obligatorio)"),
        ("Preguntas en cuestionario actual", "25"),
        ("", ""),
        ("Estado de cada pregunta", ""),
        ("→ Mantenidas tal cual", mantenidas),
        ("→ Reformuladas / Reformateadas", reformuladas),
        ("→ Movidas a otra sección", movidas),
        ("→ Fusionadas con otras", fusionadas),
        ("→ Eliminadas", eliminadas + 1),  # +1 por nombre
        ("→ Nuevas (no existían en Javier)", nuevas),
        ("", ""),
        ("Comparación de tiempo", ""),
        ("Javier (Forms): tiempo declarado / real", "10 min / 10-12 min"),
        ("Actual: tiempo declarado / real", "10 min / 6-7 min"),
    ]

    for r, (m, v) in enumerate(resumen_rows, 1):
        cell_m = ws2.cell(row=r, column=1, value=m)
        cell_v = ws2.cell(row=r, column=2, value=v)
        if r == 1:
            cell_m.font = bold_font
            cell_v.font = bold_font
            cell_m.fill = header_fill
            cell_v.fill = header_fill
            cell_m.alignment = center
            cell_v.alignment = center
        elif m and not v and v != 0:  # subtítulo
            cell_m.font = Font(bold=True, size=11, color="1E3A5F")

    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 35

    # ===== HOJA 3: LEYENDA DE COLORES =====
    ws3 = wb.create_sheet("Leyenda")
    leyenda = [
        ("Estado", "Color", "Significado"),
        ("MANTENIDA", "blanco", "Pregunta de Javier que se mantiene tal cual (texto y opciones idénticos)."),
        ("REFORMATEADA", "amarillo", "Pregunta de Javier que se mantiene pero cambia de FORMATO (ej: radio buttons a select desplegable)."),
        ("REFORMULADA", "amarillo", "Pregunta de Javier cuyo TEXTO o INTENCIÓN se ajustó (ej: a afirmación, a escala distinta)."),
        ("MOVIDA", "azul", "Pregunta movida de sección sin cambiar el contenido."),
        ("FUSIONADA", "verde", "Pregunta consolidada con otras en una nueva pregunta."),
        ("ELIMINADA", "rosa", "Pregunta de Javier que se quitó del cuestionario."),
        ("NUEVA", "morado", "Pregunta nueva que no estaba en el cuestionario original de Javier."),
    ]
    for r, row in enumerate(leyenda, 1):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.alignment = left_wrap
            cell.border = border
            if r == 1:
                cell.font = bold_font
                cell.fill = header_fill
            else:
                estado_color = ESTADOS.get(row[0], "FFFFFF")
                cell.fill = PatternFill(start_color=estado_color, end_color=estado_color, fill_type="solid")

    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 70
    ws3.row_dimensions[1].height = 28

    output = "Comparativo_Cuestionario_Moderniza.xlsx"
    wb.save(output)
    print(f"Excel generado: {output}")


if __name__ == "__main__":
    main()
