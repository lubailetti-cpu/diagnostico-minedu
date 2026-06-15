"""
Genera Excel con el cuestionario completo del diagnostico Moderniza,
mostrando la bifurcacion entre MINEDU sede central, DRE y UGEL.

4 hojas:
- Maestro: todas las preguntas + columna "Aplica a"
- Flujo MINEDU: solo las que ve un usuario MINEDU sede central
- Flujo DRE: solo las que ve un usuario DRE
- Flujo UGEL: solo las que ve un usuario UGEL
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colores por nivel
COLOR_MINEDU = "DBEAFE"
COLOR_DRE = "FEF3C7"
COLOR_UGEL = "FCE7F3"
COLOR_TODOS = "F0FDF4"
COLOR_SEC = "EFF4FF"

# Estructura: (seccion, pregunta, tipo, opciones, aplica_a, condicional)
# aplica_a: "Todos" | "MINEDU" | "DRE" | "UGEL" | "DRE/UGEL"
PREGUNTAS = [
    # === SECCIÓN A: PERFIL ===
    ("A. Perfil del servidor",
     "¿En qué nivel del sector educativo trabaja?",
     "Select desplegable",
     "MINEDU sede central / Dirección Regional de Educación (DRE) / Unidad de Gestión Educativa Local (UGEL)",
     "Todos",
     "Pregunta FILTRO. Dispara la bifurcación condicional del cuestionario."),

    ("A. Perfil del servidor",
     "¿A qué régimen laboral pertenece?",
     "Select desplegable",
     "D.L. 1057 CAS / D.L. 276 / D.L. 728 / Ley 30057 / Locador / Practicante",
     "Todos",
     ""),

    ("A. Perfil del servidor",
     "¿En qué categoría se ubica su unidad organizacional?",
     "Select desplegable (cascada)",
     "Alta Dirección / Órganos de Control y Defensa / Órganos de Asesoramiento / Órganos de Apoyo / "
     "Órganos de Línea — Gestión Pedagógica / Órganos de Línea — Gestión Institucional / Organismos Adscritos",
     "MINEDU",
     "Solo aparece si nivel = MINEDU. Basado en ROF DS 001-2015-MINEDU."),

    ("A. Perfil del servidor",
     "¿Cuál unidad específicamente?",
     "Select desplegable (cascada filtrada)",
     "Lista filtrada según la categoría elegida arriba (entre 1 y 26 opciones)",
     "MINEDU",
     "Solo aparece si nivel = MINEDU. DEBEDSAR Sede Central y COAR siguen separadas (como Javier las puso)."),

    ("A. Perfil del servidor",
     "¿A qué Dirección Regional de Educación (DRE) pertenece?",
     "Select desplegable",
     "Las 26 DRE del Perú: Amazonas, Áncash, Apurímac, Arequipa, Ayacucho, Cajamarca, Callao, Cusco, "
     "Huancavelica, Huánuco, Ica, Junín, La Libertad, Lambayeque, Lima Metropolitana, Lima Provincias, "
     "Loreto, Madre de Dios, Moquegua, Pasco, Piura, Puno, San Martín, Tacna, Tumbes, Ucayali",
     "DRE",
     "Solo aparece si nivel = DRE."),

    ("A. Perfil del servidor",
     "¿A qué UGEL pertenece?",
     "Campo de texto libre",
     "Texto libre (hay 220+ UGELs en Perú)",
     "UGEL",
     "Solo aparece si nivel = UGEL."),

    ("A. Perfil del servidor",
     "¿Cuál es el rol principal que ejerce actualmente?",
     "Select desplegable",
     "Asesor/a / Coordinador/a / Especialista técnico/a / Especialista administrativo/a / Otras",
     "Todos",
     ""),

    ("A. Perfil del servidor",
     "¿Cuántos años de experiencia tiene en el sector público?",
     "Select desplegable",
     "Menos de 2 años / De 2 a 5 / De 6 a 10 / Más de 10",
     "Todos",
     ""),

    # === SECCIÓN B: TU TRABAJO DOCUMENTAL ===
    ("B. Tu trabajo documental",
     "Distribución de su tiempo laboral semanal",
     "Barra segmentada interactiva (3 puntos arrastrables, 4 colores)",
     "4 categorías que deben sumar 100%: Funciones sustantivas / Documentos normativos e instrumentos de gestión / "
     "Tareas administrativas repetitivas / Atención de demandas externas no planificadas",
     "Todos",
     "Si no mueve los puntos al confirmar, aparece popup que pregunta si realmente las 4 actividades ocupan el mismo tiempo."),

    ("B. Tu trabajo documental",
     "Para cada tipo de documento: ¿cuántos elabora/revisa al mes y cuánto tiempo le toma cada uno?",
     "Matriz por tipo (cards verticales con pills)",
     "9 tipos: Ficha de proceso/producto, Ficha de procedimiento, Diagrama BPMN, Informe técnico, "
     "Presentación institucional, Ayuda memoria, Oficio/memorando, Documento normativo, Guía metodológica. "
     "Para cada uno: Cantidad (0/1-2/3-5/6-10/11+) y Tiempo (<30 min/30 min-2h/2-4h/4-8h/1-3 días/3+ días)",
     "Todos",
     "NÚCLEO DE LA LÍNEA DE BASE. Permite medir impacto por tipo de documento después del piloto."),

    ("B. Tu trabajo documental",
     "¿Cuál es la principal dificultad al elaborar estos instrumentos?",
     "Checkboxes (max 2)",
     "Normativa dispersa / Faltan plantillas / Muchas revisiones / Falta de tiempo / Falta de capacitación / "
     "Instrucciones poco claras / Documentos derivados incorrectamente / Otras",
     "Todos",
     ""),

    ("B. Tu trabajo documental",
     "¿Tiene acceso a plantillas, guías o herramientas estandarizadas?",
     "Opción única",
     "Sí y las uso / Sí pero desactualizadas / No tengo acceso / No las necesito",
     "Todos",
     "QUITABLE si se reduce. La matriz de volumen/tiempo ya da pista indirecta."),

    # === SECCIÓN C: REPROCESO Y CALIDAD ===
    ("C. Reproceso y calidad",
     "De cada 10 documentos que entrega, ¿cuántos le devuelven con observaciones?",
     "Opción única",
     "Casi ninguno (0-1) / Pocos (2-3) / La mitad (4-6) / Bastantes (7-8) / Casi todos (9-10)",
     "Todos",
     ""),

    ("C. Reproceso y calidad",
     "Cuando un documento requiere reproceso, ¿cuántas rondas de revisión se producen?",
     "Opción única",
     "1 ronda / 2 / 3 / Más de 3",
     "Todos",
     "QUITABLE si se reduce. El % de devolución ya da una métrica de retrabajo."),

    ("C. Reproceso y calidad",
     "Cuando le devuelven un documento, ¿por qué suele ser?",
     "Checkboxes (max 2)",
     "Errores de forma (formato, redacción, ortografía) / Errores de fondo (criterio técnico) / "
     "Falta de alineamiento normativo / Estilo institucional inadecuado / Información incompleta / Otras",
     "Todos",
     "Distingue entre errores de FORMA (que Moderniza ataca directo) y errores de FONDO (que no). Crítico para atribuir impacto."),

    # === SECCIÓN D: BÚSQUEDA DE INFORMACIÓN ===
    ("D. Búsqueda de información",
     "¿Cuánto tiempo dedica a buscar normativa, guías o modelos ANTES de elaborar un instrumento?",
     "Opción única",
     "<30 min / 30 min-1h / 1-2h / >2h / No suelo buscar, trabajo de memoria",
     "Todos",
     ""),

    ("D. Búsqueda de información",
     "¿Qué tan trabajoso le resulta asegurarse que la normativa que está usando es la versión vigente?",
     "Escala 1-4 semáforo",
     "1 (verde) = Muy fácil / 4 (rojo) = Muy trabajoso",
     "Todos",
     "DIMENSIONA EL PROBLEMA CENTRAL del RAG. Si la mayoría responde alto, justifica con datos el esfuerzo."),

    # === SECCIÓN E: RITMO Y DEMANDAS EXTERNAS ===
    ("E. Ritmo y demandas externas",
     "¿Con qué frecuencia recibe solicitudes urgentes de entidades externas (Congreso, Contraloría, PCM, etc.)?",
     "Opción única",
     "Nunca / 1-2 por semana / 3-5 por semana / Diariamente o más",
     "Todos",
     "QUITABLE si se reduce."),

    ("E. Ritmo y demandas externas",
     "Cuando recibe estas solicitudes externas, ¿cuál es el plazo habitual?",
     "Opción única",
     "Más de 3 días / Entre 1 y 3 días / Menos de 24 h / Menos de 4 h / No aplica",
     "Todos",
     "QUITABLE si se reduce. Solo importa si dejamos la pregunta anterior."),

    # === SECCIÓN F: OPORTUNIDADES EN TU ÁREA ===
    ("F. Oportunidades en tu área",
     "Valore qué tan identificado se siente con 8 afirmaciones (todas en negativo, escala 1-4 semáforo)",
     "8 afirmaciones, cada una escala 1-4 semáforo (1=verde, 4=rojo)",
     "1. La carga documental ha ocasionado retrasos en POI / "
     "2. La carga de trabajo de mi equipo NO está bien distribuida / "
     "3. Me siento agotado al cerrar la semana / "
     "4. Poco satisfecho con la calidad de mis entregables / "
     "5. Hay mucho trabajo manual repetitivo automatizable / "
     "6. Información importante en cabeza de pocas personas / "
     "7. Datos dispersos que cuesta consolidar / "
     "8. Cuesta transformar info en indicadores",
     "Todos",
     "MAPEA OPORTUNIDADES BCP. Identifica el apetito de cada área para Moderniza y futuros workflows."),

    ("F. Oportunidades en tu área",
     "En el último mes, ¿cuántas veces extendió su jornada laboral para cumplir con entregables?",
     "Opción única",
     "Nunca / 1-2 veces / 3-5 / Más de 5",
     "Todos",
     "QUITABLE si se reduce. La afirmación de agotamiento del bloque BCP ya cubre."),

    ("F. Oportunidades en tu área",
     "Si pudiera recuperar tiempo de las tareas repetitivas, ¿a qué actividades lo destinaría?",
     "Checkboxes (max 2)",
     "Análisis y diagnóstico / Mejora de procesos e innovación / Coordinación con otras unidades / "
     "Atención a ciudadanos / Capacitación y desarrollo / Otras",
     "Todos",
     "QUITABLE si se reduce."),

    # === SECCIÓN G: USO ACTUAL DE IA ===
    ("G. Uso actual de IA",
     "¿Cuál es tu nivel actual de uso de inteligencia artificial?",
     "Opción única",
     "No uso / Nivel 1 Básico (consultas puntuales) / Nivel 2 Pragmático (resumir, redactar) / "
     "Nivel 3 Constructor (automatizo tareas) / Nivel 4 Avanzado (agentes integrados)",
     "Todos",
     "Escala BCP de madurez. Identifica early adopters."),

    ("G. Uso actual de IA",
     "¿Usas alguna herramienta de IA para facilitar tu trabajo?",
     "Opción única",
     "Sí, con cuenta pagada por el MINEDU / Sí, con cuenta pagada por mí / Sí, con cuenta gratuita / No uso ninguna",
     "Todos",
     ""),

    ("G. Uso actual de IA",
     "¿Para qué tareas de tu trabajo usas IA hoy?",
     "Checkboxes (max 3)",
     "Redactar borradores / Resumir documentos / Buscar información / Traducir / Generar ideas / "
     "Analizar datos Excel / Hacer presentaciones / Otra / No uso IA para mi trabajo",
     "Todos",
     "CONDICIONAL: solo aparece si en la anterior respondió Sí (alguna opción de pagada o gratuita)."),

    # === SECCIÓN H: PERCEPCIÓN PLATAFORMA ===
    ("H. Percepción sobre plataforma IA",
     "Valore 4 afirmaciones sobre una posible plataforma de asistencia con IA",
     "4 afirmaciones, escala 1-5",
     "1. Reduciría tiempo en tareas repetitivas / "
     "2. Ayudaría a generar borradores de mayor calidad / "
     "3. Daría más tiempo para actividades estratégicas / "
     "4. Me gustaría participar en una prueba piloto",
     "Todos",
     ""),

    ("H. Percepción sobre plataforma IA",
     "¿Cuáles serían sus principales preocupaciones al usar una herramienta de IA?",
     "Checkboxes (max 2)",
     "Errores técnicos o normativos / Pérdida de criterio profesional / No adaptada al MINEDU / "
     "Seguridad o confidencialidad / Resistencia institucional / Tiempo de capacitación / Sin preocupaciones / Otras",
     "Todos",
     "QUITABLE si se reduce. Útil para venta, no para impacto."),

    ("H. Percepción sobre plataforma IA",
     "¿Qué condición considera indispensable para adoptar IA en su trabajo?",
     "Checkboxes (max 2)",
     "Validación institucional / Capacitación previa / Fácil de usar / Confidencialidad garantizada / "
     "Soporte técnico disponible / Poder revisar el borrador antes",
     "Todos",
     "QUITABLE si se reduce. Útil para venta, no para impacto."),

    # === SECCIÓN I bis: DRE / UGEL ESPECÍFICO ===
    ("I. Trabajo específico DRE / UGEL",
     "¿Cuántas Instituciones Educativas (IIEE) están a tu cargo o jurisdicción?",
     "Opción única",
     "Menos de 10 / 10 a 50 / 51 a 200 / Más de 200",
     "DRE/UGEL",
     "CONDICIONAL: solo aparece si nivel = DRE o UGEL. Segmenta por escala de gestión."),

    ("I. Trabajo específico DRE / UGEL",
     "¿Cuántas visitas o supervisiones a IIEE realizas en promedio al mes?",
     "Opción única",
     "Ninguna / 1 a 3 / 4 a 10 / Más de 10",
     "DRE/UGEL",
     "Identifica intensidad del trabajo de campo."),

    ("I. Trabajo específico DRE / UGEL",
     "Después de una visita a IIEE, ¿cuánto tiempo te toma elaborar el informe?",
     "Opción única",
     "<1 hora / 1-3 horas / 3-8 horas / Más de un día",
     "DRE/UGEL",
     "ALTO POTENCIAL: workflow Generador de Informes de Supervisión a partir de notas de campo o audio."),

    ("I. Trabajo específico DRE / UGEL",
     "Antes de una visita, ¿cuánto tiempo dedicas a preparar antecedentes?",
     "Opción única",
     "<30 min / 30 min-2 h / >2 h / No preparo antecedentes",
     "DRE/UGEL",
     "ALTO POTENCIAL: workflow Pre-visita que consolida historial de IIEE en brief automático."),

    ("I. Trabajo específico DRE / UGEL",
     "¿Cuánto tiempo dedicas a consolidar respuestas de IIEE en formatos distintos?",
     "Opción única",
     "<2 h por semana / 2-5 h / 5-10 h / Más de 10 h por semana",
     "DRE/UGEL",
     "ALTO POTENCIAL: workflow Consolidador que estandariza respuestas heterogéneas en una sola matriz."),

    ("I. Trabajo específico DRE / UGEL",
     "¿Cuántas normas o directivas del MINEDU traduces o adaptas para las IIEE al mes?",
     "Opción única",
     "Ninguna / 1-3 / 4-10 / Más de 10",
     "DRE/UGEL",
     "ALTO POTENCIAL: workflow Traductor de Normas (norma → flyer + comunicado claro para directores)."),

    ("I. Trabajo específico DRE / UGEL",
     "¿Cuánto tiempo dedicas a responder consultas frecuentes de directores y docentes?",
     "Opción única",
     "<2 h por semana / 2-5 h / 5-10 h / Más de 10 h por semana",
     "DRE/UGEL",
     "ALTO POTENCIAL: chatbot interno para consultas frecuentes (Mo pero para DRE/UGEL)."),

    ("I. Trabajo específico DRE / UGEL",
     "¿Para cuáles tareas de tu unidad sería MÁS útil una herramienta con IA?",
     "Checkboxes (max 3)",
     "Preparar visitas / Elaborar informes post-visita / Consolidar reportes IIEE / "
     "Traducir normas a directores / Responder consultas frecuentes / Generar oficios masivos / "
     "Procesar trámites docentes / Capacitaciones / Otra",
     "DRE/UGEL",
     "MAPEO DIRECTO de oportunidades de workflow. Es la pregunta más valiosa para el roadmap DRE/UGEL."),

    # === SECCIÓN J: TU VOZ CUENTA ===
    ("J. Tu voz cuenta",
     "En sus propias palabras: ¿cuáles son las 2 o 3 tareas repetitivas que más tiempo le quitan?",
     "Texto libre",
     "Texto abierto",
     "Todos",
     "Pregunta cualitativa. Capta dolores no listados en las opciones cerradas."),
]


def color_aplica(aplica):
    if aplica == "Todos":
        return COLOR_TODOS
    elif aplica == "MINEDU":
        return COLOR_MINEDU
    elif aplica == "DRE":
        return COLOR_DRE
    elif aplica == "UGEL":
        return COLOR_UGEL
    elif aplica == "DRE/UGEL":
        return COLOR_DRE
    return "FFFFFF"


def should_include(aplica, perfil):
    """Decide si una pregunta debe incluirse en el flujo de un perfil."""
    if aplica == "Todos":
        return True
    if perfil == "MINEDU" and aplica == "MINEDU":
        return True
    if perfil == "DRE" and aplica in ("DRE", "DRE/UGEL"):
        return True
    if perfil == "UGEL" and aplica in ("UGEL", "DRE/UGEL"):
        return True
    return False


def write_sheet(wb, name, perfil=None):
    ws = wb.create_sheet(name) if name != "Maestro" else wb.active
    if name == "Maestro":
        ws.title = "Maestro"

    bold_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin = Side(border_style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    if perfil:
        titulo = f"Flujo del cuestionario para usuarios: {perfil}"
    else:
        titulo = "Cuestionario MAESTRO — todas las preguntas con bifurcación"

    cell_t = ws.cell(row=1, column=1, value=titulo)
    cell_t.font = Font(name='Calibri', size=14, bold=True, color="1E3A5F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7 if name == "Maestro" else 6)

    # Subtítulo (descripción)
    if perfil:
        subt = f"Esta hoja muestra solo las preguntas que ve un usuario de tipo {perfil}. Total estimado: ver al final."
    else:
        subt = ("Esta hoja muestra todas las preguntas posibles del cuestionario. La columna 'Aplica a' indica qué tipo "
                "de usuario las ve. Pestañas adjuntas para ver solo el flujo de cada perfil.")
    cell_st = ws.cell(row=2, column=1, value=subt)
    cell_st.font = Font(name='Calibri', size=10, italic=True, color="5C6178")
    cell_st.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7 if name == "Maestro" else 6)
    ws.row_dimensions[2].height = 40

    # Headers
    if name == "Maestro":
        headers = ["#", "Sección", "Pregunta", "Tipo de respuesta", "Opciones / Escala", "Aplica a", "Comentario"]
    else:
        headers = ["#", "Sección", "Pregunta", "Tipo de respuesta", "Opciones / Escala", "Comentario"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 32

    # Filas
    row_idx = 5
    n = 0
    last_sec = ""
    for sec, preg, tipo, ops, aplica, com in PREGUNTAS:
        if perfil and not should_include(aplica, perfil):
            continue
        n += 1
        # Color de fila según aplica_a
        row_color = color_aplica(aplica) if name == "Maestro" else "FFFFFF"

        # Si cambia la sección, separador visual
        if sec != last_sec and name != "Maestro":
            ws.cell(row=row_idx, column=1, value=sec).font = Font(bold=True, size=11, color="1E3A5F")
            ws.cell(row=row_idx, column=1).fill = PatternFill(start_color=COLOR_SEC, end_color=COLOR_SEC, fill_type="solid")
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1
            last_sec = sec

        if name == "Maestro":
            vals = [n, sec, preg, tipo, ops, aplica, com]
        else:
            vals = [n, sec, preg, tipo, ops, com]

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.alignment = left_wrap
            cell.border = border
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            if name == "Maestro" and col == 6:  # columna Aplica a
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.alignment = center
        ws.row_dimensions[row_idx].height = 70
        row_idx += 1

    # Anchos
    if name == "Maestro":
        widths = [6, 26, 50, 26, 60, 14, 55]
    else:
        widths = [6, 26, 50, 26, 60, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    # Total al final
    total_row = ws.cell(row=row_idx + 1, column=1, value=f"Total preguntas: {n}")
    total_row.font = Font(name='Calibri', size=11, bold=True, color="1E3A5F")


def main():
    wb = openpyxl.Workbook()
    write_sheet(wb, "Maestro")
    write_sheet(wb, "Flujo MINEDU central", perfil="MINEDU")
    write_sheet(wb, "Flujo DRE", perfil="DRE")
    write_sheet(wb, "Flujo UGEL", perfil="UGEL")

    # ===== HOJA "Resumen visual" =====
    ws_res = wb.create_sheet("Resumen visual")
    bold_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    rows = [
        ("Perfil de usuario", "Preguntas que ve", "Tiempo estimado", "Bloques específicos"),
        ("MINEDU sede central",
         str(sum(1 for p in PREGUNTAS if should_include(p[4], "MINEDU"))),
         "~7-8 minutos",
         "Cascada categoría + unidad MINEDU (ROF). NO ve el bloque DRE/UGEL ni los selectores de DRE/UGEL."),
        ("DRE (Dirección Regional)",
         str(sum(1 for p in PREGUNTAS if should_include(p[4], "DRE"))),
         "~9-10 minutos",
         "Select de DRE (26 regiones). Bloque específico de 8 preguntas DRE/UGEL al final. NO ve cascada MINEDU ni UGEL."),
        ("UGEL (Unidad Gestión Educativa Local)",
         str(sum(1 for p in PREGUNTAS if should_include(p[4], "UGEL"))),
         "~9-10 minutos",
         "Input texto para UGEL. Bloque específico de 8 preguntas DRE/UGEL al final. NO ve cascada MINEDU ni DRE."),
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws_res.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if r == 1:
                cell.font = bold_font
                cell.fill = header_fill
            else:
                cell.font = Font(name='Calibri', size=11)
                if c == 1:
                    if "MINEDU" in row[0]:
                        cell.fill = PatternFill(start_color=COLOR_MINEDU, end_color=COLOR_MINEDU, fill_type="solid")
                    elif "DRE" in row[0]:
                        cell.fill = PatternFill(start_color=COLOR_DRE, end_color=COLOR_DRE, fill_type="solid")
                    elif "UGEL" in row[0]:
                        cell.fill = PatternFill(start_color=COLOR_UGEL, end_color=COLOR_UGEL, fill_type="solid")
        ws_res.row_dimensions[r].height = 50 if r > 1 else 28

    ws_res.column_dimensions['A'].width = 35
    ws_res.column_dimensions['B'].width = 18
    ws_res.column_dimensions['C'].width = 20
    ws_res.column_dimensions['D'].width = 70

    output = "Cuestionario_Completo_Moderniza.xlsx"
    wb.save(output)
    print(f"Excel generado: {output}")
    print(f"Total preguntas en el cuestionario (maestro): {len(PREGUNTAS)}")
    print(f"  - MINEDU central ve: {sum(1 for p in PREGUNTAS if should_include(p[4], 'MINEDU'))}")
    print(f"  - DRE ve: {sum(1 for p in PREGUNTAS if should_include(p[4], 'DRE'))}")
    print(f"  - UGEL ve: {sum(1 for p in PREGUNTAS if should_include(p[4], 'UGEL'))}")


if __name__ == "__main__":
    main()
